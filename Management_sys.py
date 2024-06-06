import tkinter as tk
from tkinter import ttk, messagebox
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font
import os
import time
import threading

class General(tk.Tk):
    def __init__(self):
        super().__init__()
        self.title('Management System Software')
        self.geometry('1350x750')

        # Form widgets
        header_font = ("Cambria", 13, "bold")
        text_font = ("Times", 11)

        self.entries = {}
        self.comboboxes = {}
        self.radio_vars = {}

        self.create_form(header_font, text_font)
        self.create_buttons()

    def create_form(self, header_font, text_font):
        labels = [
            "Name", "Email Id", "Contact No.", "City", "State", "College", 
            "Branch", "Semester", "Year Of Passing", "Course", "Batch", 
            "Training Center", "SEX", "Save File As", "Query/Regarding What"
        ]

        row = 0
        for label in labels:
            lbl = tk.Label(self, text=f"{label}: ", font=header_font)
            lbl.grid(row=row, column=0, pady=10, padx=10, sticky="e")

            if label in ["State", "Semester", "Year Of Passing", "Course", "Training Center"]:
                combo = ttk.Combobox(self, font=text_font)
                if label == "State":
                    combo['values'] = [
                        'SELECT', 'Andhra Pradesh', 'Arunachal Pradesh', 'Assam', 'Bihar', 'Goa', 
                        'Gujarat', 'Haryana', 'Himachal Pradesh', 'Jammu & Kashmir', 'Karnataka', 
                        'Kerala', 'Madhya Pradesh', 'Maharashtra', 'Manipur', 'Meghalaya', 
                        'Mizoram', 'Nagaland', 'Orissa', 'Punjab', 'Rajasthan', 'Sikkim', 
                        'Tamil Nadu', 'Tripura', 'Uttar Pradesh', 'West Bengal', 'Chhattisgarh', 
                        'Uttarakhand', 'Jharkhand', 'Telangana'
                    ]
                elif label == "Semester":
                    combo['values'] = [str(i) for i in range(1, 9)] + ["Passed Out"]
                elif label == "Year Of Passing":
                    combo['values'] = ["SELECT"] + [str(i) for i in range(2050, 1999, -1)]
                elif label == "Course":
                    combo['values'] = [
                        'SELECT COURSE', 'ESR (30 Days)', 'ESR (45 Days)', 'Matlab (30 Days)', 
                        'IOT (15 Days)', 'IOT (30 Days)', 'JAVA (30 Days)', 'Python (30 Days)', 
                        'PLC-SCADA (30 Days)', 'C/C++ (45 Days)', 'Android (15 Days)', 'Android (30 Days)'
                    ]
                elif label == "Training Center":
                    combo['values'] = [
                        'SELECT CENTER', 'Jaipur', 'Hyderabad', 'Raipur', 'Lucknow', 'Pune', 
                        'Vizag', 'Bhopal', 'Delhi'
                    ]
                combo.current(0)
                combo.grid(row=row, column=1, pady=10, padx=10, sticky="w")
                self.comboboxes[label] = combo
            elif label == "SEX":
                var = tk.StringVar()
                frame = tk.Frame(self)
                frame.grid(row=row, column=1, pady=10, padx=10, sticky="w")
                for option in ["Male", "Female"]:
                    radio = tk.Radiobutton(frame, text=option, variable=var, value=option, font=("Calibri", 10, "bold"))
                    radio.pack(side="left")
                self.radio_vars[label] = var
            else:
                entry = tk.Entry(self, font=text_font)
                entry.grid(row=row, column=1, pady=10, padx=10, sticky="w")
                if label == "Save File As":
                    entry.insert(0, "Student_List.xlsx")
                self.entries[label] = entry
            row += 1

        # Training Session Radio Buttons
        lbl = tk.Label(self, text="Training Session: ", font=header_font)
        lbl.grid(row=row, column=0, pady=10, padx=10, sticky="e")

        var = tk.StringVar()
        frame = tk.Frame(self)
        frame.grid(row=row, column=1, pady=10, padx=10, sticky="w")
        for option in ["Summer Training", "Winter Training", "Project Based", "Other"]:
            radio = tk.Radiobutton(frame, text=option, variable=var, value=option, font=("Calibri", 10, "bold"))
            radio.pack(side="left")
        self.radio_vars["Training Session"] = var

    def create_buttons(self):
        save_btn = tk.Button(self, text="SAVE", font=("Calibri", 13), command=self.save_clicked)
        save_btn.grid(row=16, column=0, pady=20)

        clear_btn = tk.Button(self, text="CLEAR", font=("Calibri", 13), command=self.clear_clicked)
        clear_btn.grid(row=16, column=1, pady=20)

    def valid_contact(self, phone_number):
        return phone_number.isdigit() and len(phone_number) == 10

    def savesuccess(self):
        messagebox.showinfo("SAVED", "Your entries have been saved successfully!")

    def errorcontactmsg(self):
        messagebox.showwarning("Invalid Contact Number", "Please enter a valid contact number!")

    def erroremailmsg(self):
        messagebox.showwarning("Invalid Email ID", "Please enter a valid Email ID!")

    def errornormmsg(self):
        messagebox.showwarning("Error Report", "Please enter the necessary fields: Name, Email Id, Contact No.!!")

    def crashingmsg(self):
        messagebox.showerror("Program Crashing", "The XML file is already open. Please close the XML file and try again.")

    def save_clicked(self):
        try:
            tex = {label: entry.get() for label, entry in self.entries.items()}
            combobox_vals = {label: combo.get() for label, combo in self.comboboxes.items()}
            radio_vals = {label: var.get() for label, var in self.radio_vars.items()}

            if tex["Name"] and tex["Email Id"] and tex["Contact No."]:
                if '@' in tex["Email Id"]:
                    if self.valid_contact(tex["Contact No."]):
                        save_file = tex["Save File As"]
                        try:
                            wb = load_workbook(save_file)
                            ws = wb.active
                            row = ws.max_row + 1
                        except FileNotFoundError:
                            wb = Workbook()
                            ws = wb.active
                            ws.title = "Student_List"
                            ft = Font(bold=True)
                            headers = [
                                "S.NO", "TOKEN NO.", "NAME", "EMAIL ID", "CONTACT NO.", "CITY", "STATE", "COLLEGE", 
                                "BRANCH", "SEMESTER", "YEAR OF PASSING", "COURSE", "DATE OF REGISTERATION", "BATCH", 
                                "TRAINING CENTER", "TRAINING SESSION", "SEX", "QUERY/REGARDING WHAT", "TOTAL FEE", 
                                "SUBMITTED FEE", "REMAINING FEE", "DISCOUNT", "REMARKS", "ADDITIONAL REMARKS", 
                                "SYSTEM DATE AND TIME"
                            ]
                            ws.append(headers)
                            for col in ws.iter_cols(min_row=1, max_row=1, min_col=1, max_col=len(headers)):
                                for cell in col:
                                    cell.font = ft
                            row = 2

                        sno = row - 1
                        tkno = row - 1
                        data = [
                            sno, tkno, tex["Name"], tex["Email Id"], tex["Contact No."], tex["City"], 
                            combobox_vals["State"], tex["College"], tex["Branch"], combobox_vals["Semester"], 
                            combobox_vals["Year Of Passing"], combobox_vals["Course"], time.strftime("%Y-%m-%d"), 
                            tex["Batch"], combobox_vals["Training Center"], radio_vals["Training Session"], 
                            radio_vals["SEX"], tex["Query/Regarding What"], "", "", "", "", "", "", 
                            time.strftime("%Y-%m-%d %H:%M:%S")
                        ]
                        ws.append(data)
                        wb.save(save_file)
                        threading.Thread(target=self.savesuccess).start()
                    else:
                        threading.Thread(target=self.errorcontactmsg).start()
                else:
                    threading.Thread(target=self.erroremailmsg).start()
            else:
                threading.Thread(target=self.errornormmsg).start()
        except PermissionError:
            threading.Thread(target=self.crashingmsg).start()

    def clear_clicked(self):
        for entry in self.entries.values():
            entry.delete(0, tk.END)
        for combo in self.comboboxes.values():
            combo.current(0)
        for var in self.radio_vars.values():
            var.set("")

if __name__ == "__main__":
    app = General()
    app.mainloop()
